import sys
import os
import requests
import json
import re
import time  # 추가된 부분
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from urllib.parse import urljoin, urlparse
import google.generativeai as genai
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font
from PIL import Image
from io import BytesIO
from tqdm import tqdm
from openpyxl.drawing.image import Image as OpenpyxlImage

import logging
from logging.handlers import RotatingFileHandler


# 로그 설정
def setup_logging():
    # 기존 로거가 있다면 중복 방지
    logger = logging.getLogger("ProductScraper")
    if logger.hasHandlers():
        return logger  # 이미 설정된 로거 반환

    logger.setLevel(logging.DEBUG)  # 로그 레벨 설정

    # 로그 포맷 설정
    formatter = logging.Formatter(
        '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )

    # 콘솔 핸들러
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)  # 콘솔에서는 INFO 이상만 출력
    console_handler.setFormatter(formatter)

    # 파일 핸들러
    file_handler = RotatingFileHandler(
        "scraper.log", maxBytes=5 * 1024 * 1024, backupCount=3
    )
    file_handler.setLevel(logging.DEBUG)  # 파일에는 DEBUG 이상 모두 기록
    file_handler.setFormatter(formatter)

    # 핸들러 추가
    logger.addHandler(console_handler)
    logger.addHandler(file_handler)

    return logger


# 로거 초기화
logger = setup_logging()



def get_resource_path(relative_path):
    """PyInstaller 빌드 환경에서 리소스 파일 경로를 반환."""
    try:
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# 설정 파일 로드
config_path = get_resource_path("config.json")
category_urls_path = get_resource_path("category_urls.txt")



# 설정 파일 로드
with open(config_path, 'r', encoding='utf-8') as file:
    config_data = json.load(file)
    api_key = config_data['api_key']
    model_name = config_data['model']

genai.configure(api_key=api_key)
model = genai.GenerativeModel(
    model_name=model_name,
    generation_config={
        "temperature": 0,
        "top_p": 0.95,
        "top_k": 64,
        "max_output_tokens": 500,
        "response_mime_type": "application/json",
    }
)

# 카테고리 URL 읽기
with open(category_urls_path, 'r', encoding='utf-8') as file:
    category_data = [line.strip().split('|') for line in file.readlines()]

results = {}
timestamp = datetime.now().strftime("%Y%m%d%H%M%S")

def setup_driver():
    chrome_options = Options()
    # chrome_options.add_argument("--headless=new")  # 최신 Headless 모드
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--disable-extensions")
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    chrome_options.add_argument("--disable-infobars")
    chrome_options.add_argument("--ignore-certificate-errors")  # SSL 인증 무시
    chrome_options.add_argument("--ignore-ssl-errors=yes")      # SSL 오류 무시
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option('useAutomationExtension', False)

    chrome_options.add_argument("--ssl-protocol=any")
    chrome_options.add_argument("--tlsv1")
    chrome_options.add_argument("--tlsv1.1")
    chrome_options.add_argument("--tlsv1.2")


    # User-Agent 설정
    chrome_options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/107.0.0.0 Safari/537.36"
    )

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=chrome_options)
    return driver

def is_valid_image(img_content):
    try:
        img = Image.open(BytesIO(img_content))
        img.verify()
        if img.height < 200:
            return False
        return True
    except (IOError, SyntaxError):
        return False

def get_product_urls(category_url, site_name):
    driver = setup_driver()
    product_urls = set()  # 중복 제거를 위해 set 사용
    
    try:
        if '퀄엔드' in site_name:
            driver.get(category_url)
            time.sleep(5)
            
            # 페이지 끝까지 스크롤
            last_height = driver.execute_script("return document.body.scrollHeight")
            while True:
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                time.sleep(2)
                new_height = driver.execute_script("return document.body.scrollHeight")
                if new_height == last_height:
                    break
                last_height = new_height
            
            html_content = driver.page_source
            soup = BeautifulSoup(html_content, 'html.parser')
            
            # 상품 링크 찾기 - 더 구체적인 selector 사용
            product_containers = soup.select('div.col-sm-3')
            
            for container in product_containers:
                link = container.find('a', href=lambda x: x and 'it_id' in x)
                if link:
                    href = link.get('href')
                    full_url = urljoin(category_url, href)
                    product_urls.add(full_url)  # set을 사용하여 중복 제거
            
            print(f"총 상품 수: {len(product_urls)}")

        elif '네임밸류' in site_name:
            # 다른 사이트는 나중에 처리
            pass
            
        elif '바이헤븐' in site_name:
            # 다른 사이트는 나중에 처리
            pass
    
    finally:
        driver.quit()
    
    urls_list = list(product_urls)
    print(f"\n중복 제거 후 URL 수: {len(urls_list)}")
    return urls_list[:100]



def ai_parse(html_data):
    prompt = f"""```html_data
{html_data}
```

```available_brand_data
[
    "ASK YOURSELF",
    "ACNE STUDIOS",
    "ALEXANDER MCQUEEN",
    "ALEXANDER WANG",
    "ALYX",
    "AMI",
    "AMIRI",
    "ARCTERYX",
    "AUDEMARS PIGUET",
    "BALENCIAGA",
    "BALMAIN",
    "BAPE",
    "BERLUTI",
    "BLANCPAIN",
    "BOTTEGA VENETA",
    "BREGUET",
    "BALLY",
    "BREITLING",
    "BRUNELLO CUCINELLI",
    "BULGARI",
    "BURBERRY",
    "CANADA GOOSE",
    "CARTIER",
    "CASABLANCA",
    "CELINE",
    "CHANEL",
    "CHAUMET",
    "CHLOE",
    "CHROME HEARTS",
    "COMME DES GARCONS",
    "CP COMPANY",
    "DELVAUX",
    "DRIES VAN NOTEN",
    "DIESEL",
    "DIOR",
    "DOLCE & GABBANA",
    "EMPORIO ARMANI",
    "FEAR OF GOD",
    "FENDI",
    "FERRAGAMO",
    "GALLERY DEPT",
    "GENTLE MONSTER",
    "GIVENCHY",
    "GOLDEN GOOSE",
    "GOYARD",
    "GUCCI",
    "HERMES",
    "HUBLOT",
    "ISABEL MARANT",
    "IAB STUDIO",
    "IWC",
    "JACQUEMUS",
    "JIL SANDER",
    "JUNJI",
    "JIMMY CHOO",
    "JORDAN",
    "JUNYA WATANABE",
    "KENZO",
    "LANVIN BLANC",
    "LANVIN",
    "LEMAIRE",
    "LOEWE",
    "LORO PIANA",
    "LOUBOUTIN",
    "LOUIS VUITTON",
    "MACKAGE",
    "MAISON MARGIELA",
    "MAISON MIHARA YASUHIRO",
    "MANOLO BLAHNIK",
    "MARNI",
    "MARTINE ROSE",
    "MAX MARA",
    "MAISON KITSUNE",
    "MIU MIU",
    "MONCLER",
    "MOOSE KNUCKLES",
    "NEW BALANCE",
    "NIKE",
    "OFF WHITE",
    "OMEGA",
    "PHILIPP PLEIN",
    "PANERAI",
    "PARAJUMPERS",
    "PALM ANGELS",
    "PALACE",
    "PATEK PHILIPPE",
    "PRADA",
    "PIAGET",
    "POLORALPHLAUREN",
    "RAY BAN",
    "RHUDE",
    "RICK OWENS",
    "RIMOWA",
    "ROGER VIVIER",
    "ROLEX",
    "SACAI",
    "SUPREME",
    "SAINT LAURENT",
    "SALOMON",
    "STUSSY",
    "STONE ISLAND",
    "TAG HEUER",
    "THE NORTH FACE",
    "THOM BROWNE",
    "TIFFANY & CO",
    "TOM FORD",
    "TUDOR",
    "UMA WANG",
    "VACHERON CONSTANTIN",
    "VALENTINO",
    "VETEMENTS",
    "VANCLEEF",
    "VERSACE",
    "WOOYOUNGMI",
    "YEEZY",
    "ZEGNA",
    "OTHERS"
    ]
```

```available_category_data
{{
    "상의": ["반팔 티셔츠", "긴팔 티셔츠", "니트/가디건", "맨투맨", "후드", "원피스", "셔츠", "드레스", "슬리브리스", "셋업", "기타 상의"],
    "아우터": ["집업", "자켓", "패딩", "레더", "코트", "기타 아우터"],
    "하의": ["팬츠", "쇼츠", "트레이닝 팬츠", "데님", "스커트", "기타 하의"],
    "가방": ["미니백", "백팩", "숄더백", "토트백", "크로스백", "클러치", "캐리어", "핸드백", "더플백", "버킷백", "기타 가방"],
    "신발": ["스니커즈", "샌들/슬리퍼", "플랫", "로퍼", "더비/레이스업", "힐/펌프스", "부츠", "기타 신발"],
    "지갑": ["반지갑", "카드지갑", "지퍼장지갑", "중/장지갑", "여권지갑", "WOC", "기타 지갑"],
    "시계": ["메탈", "가죽", "우레탄"],
    "패션잡화": ["머플러/스카프", "아이웨어", "넥타이", "모자", "헤어액세서리", "기타 잡화"],
    "액세서리": ["반지", "목걸이", "팔찌", "귀걸이", "키링", "브로치", "기타 ACC"],
    "벨트": []
}}
```

Process the given html_data into a comma-separated dict format JSON data containing the following elements.

price : int (상품의 판매 가격),
market_price : str (상품의 정품 가격 또는 매장 가격. 찾을 수 없다면 공백),
brand : string (상품의 영어 브랜드 이름. 반드시 available_brand_data 에 포함되어야 함. 포함되지 않는다면 공백),
first_category : string (상품의 1차 카테고리 분류. 반드시 available_category_data의 key 에 포함되어야 함. 포함되지 않는다면 공백),
second_category : string (상품의 1차 카테고리 분류. 반드시 available_category_data의 list에 포함되어야 함. 포함되지 않거나 first_category가 공백이라면 공백),
gender : string (상품의 대상 성별. '남성', '여성', '남성,여성' 중 하나. 정확하지 않다면 '남성,여성'),
colors : list(string) (상품의 색상 옵션값. 찾을 수 없다면 []),
sizes : list(string) (상품의 사이즈 옵션값. 찾을 수 없다면 []),
kor_name : string (상품의 한글 이름. 이름 앞에 브랜드가 딱 한번 적혀 있어야 하며 반드시 한글이어야 함),
eng_name : string (상품의 한글 이름의 영어 번역 결과. 이름 앞에 브랜드가 딱 한번 적혀 있어야 하며 반드시 영어여야 함),
genuine_number : string (상품의 정품 코드. 정품 번호는 제품 이름에 의미 없는 문자와 숫자의 조합으로 표시될 수 있음. 찾을 수 없다면 공백)
"""

    response = model.generate_content(prompt).text.strip()
    return json.loads(response)
def process_product(url, store_name, folder_name):
    driver = setup_driver()
    success = False
    
    try:
        driver.get(url)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
        html_data = driver.page_source
        
        # 이미지 처리
        folder_path = os.path.join(f"이미지/{store_name}", folder_name)
        os.makedirs(folder_path, exist_ok=True)
        
        soup = BeautifulSoup(html_data, 'html.parser')
        img_tags = soup.find_all("img")
        thumb_path = ""
        
        idx = 0
        for img in img_tags:
            if 'src' not in img.attrs:
                continue
                
            img_url = urljoin(url, img['src'])
            if (';base64,' in img_url or 
                img_url.lower().endswith('.svg') or 
                '//img.echosting.cafe24.com/' in img_url or 
                '/theme/' in img_url or 
                any(x in img_url.lower() for x in ['facebook', 'icon', 'logo', 'common', 'banner', 'brand'])):
                continue
            
            try:
                img_response = requests.get(img_url, timeout=10)
                img_response.raise_for_status()
                
                if not is_valid_image(img_response.content):
                    continue
                    
                img_path = os.path.join(folder_path, f"{idx}.jpg")
                
                with open(img_path, 'wb') as f:
                    f.write(img_response.content)
                
                if idx == 0:
                    thumb_path = img_path
                
                idx += 1
                
            except Exception:
                continue
        
        # AI 파싱
        if thumb_path:
            try:
                parsed_data = ai_parse(html_data)
                results[url].update({
                    '결과': "성공",
                    '이미지': thumb_path,
                    '매장가': parsed_data['market_price'],
                    '단가': parsed_data['price'],
                    '성별': parsed_data['gender'],
                    '상품명': re.match(r"^\[.*?\] (.*)", str(parsed_data['kor_name'])).group(1) if re.match(r"^\[.*?\] (.*)", str(parsed_data['kor_name'])) else str(parsed_data['kor_name']),
                    '영문명': re.match(r"^\[.*?\] (.*)", str(parsed_data['eng_name'])).group(1) if re.match(r"^\[.*?\] (.*)", str(parsed_data['eng_name'])) else str(parsed_data['eng_name']),
                    '브랜드': parsed_data['brand'].upper(),
                    '2차': parsed_data['first_category'],
                    '3차': parsed_data['second_category'],
                    "추가 정보\n모델명": str(parsed_data['genuine_number']),
                    "필수옵션\n색상": ",".join(parsed_data['colors']),
                    "필수옵션\n사이즈": ",".join(parsed_data['sizes']).replace("(","[").replace(")","]")
                })
                success = True
            except Exception:
                results[url]['결과'] = "실패"
    except Exception:
        results[url]['결과'] = "실패"
    finally:
        driver.quit()
        
    return success

def main():
    total_processed = 0
    
    for site_name, category_name, category_url in category_data:
        logger.info(f"\n[{site_name}] 신상품 URL 수집 중...")
        
        product_urls = get_product_urls(category_url, site_name)
        url_count = len(product_urls)
        
        print(f"[{site_name}] {url_count}개의 상품 URL 수집 완료")
        print(f"[{site_name}] 상품 정보 수집 시작...")
        
        success_count = 0
        fail_count = 0
        
        for url in tqdm(product_urls, desc=f"{site_name} 처리중"):
            folder_name = datetime.now().strftime("%Y%m%d%H%M%S")
            
            results[url] = {
                "결과": "",
                "상품넘버": (f'=HYPERLINK("{url}", "{folder_name}")'),
                "거래처": site_name,
                "단가": "",                
                "이미지": "",
                "1차": "",
                "2차": "",
                "3차": "",
                "4차": "",
                "필터": "",
                "성별": "",
                "브랜드": "",
                "2차 브랜드": "",
                "상품명": "",
                "영문명": "",                
                "추가 정보\n모델명": "",
                "추가 정보\n배송방법": "항공특송",
                "추가 정보\n소재": "",
                "추가 정보\n구성품": "풀박스",
                "매장가": "",
                "판매가1": "",
                "판매가2": "",
                "판매가3": "",
                "필수옵션\n등급선택": "",
                "필수옵션\n사이즈": "",
                "필수옵션\n색상": "",
                "필수옵션\n굽높이": "",
                "필수옵션\n버클": "",
                "필수옵션\n도금방식": "",
                "필수옵션\n밴드": "",
            }
            
            if process_product(url, site_name, folder_name):
                success_count += 1
            else:
                fail_count += 1
        
        total_processed += url_count
        print(f"\n[{site_name}] 처리 완료")
        print(f"성공: {success_count}개")
        print(f"실패: {fail_count}개")
    
    print(f"\n전체 처리 완료")
    print(f"총 처리 상품 수: {total_processed}개")
    
    return results

if __name__ == "__main__":
    import sys
    from io import StringIO

    # stdout 리셋
    sys.stdout = StringIO()

    try:
        # 리팩토링된 메인 로직 실행
        results = main()
        final = list(results.values())

        # 브랜드와 카테고리 유효성 검사
        avail_brands = [
            "ASK YOURSELF", "ACNE STUDIOS", "ALEXANDER MCQUEEN", "ALEXANDER WANG", "ALYX",
            "AMI", "AMIRI", "ARCTERYX", "AUDEMARS PIGUET", "BALENCIAGA", "BALMAIN",
            "BAPE", "BERLUTI", "BLANCPAIN", "BOTTEGA VENETA", "BREGUET", "BALLY",
            "BREITLING", "BRUNELLO CUCINELLI", "BULGARI", "BURBERRY", "CANADA GOOSE",
            "CARTIER", "CASABLANCA", "CELINE", "CHANEL", "CHAUMET", "CHLOE",
            "CHROME HEARTS", "COMME DES GARCONS", "CP COMPANY", "DELVAUX",
            "DRIES VAN NOTEN", "DIESEL", "DIOR", "DOLCE & GABBANA", "EMPORIO ARMANI",
            "FEAR OF GOD", "FENDI", "FERRAGAMO", "GALLERY DEPT", "GENTLE MONSTER",
            "GIVENCHY", "GOLDEN GOOSE", "GOYARD", "GUCCI", "HERMES", "HUBLOT",
            "ISABEL MARANT", "IAB STUDIO", "IWC", "JACQUEMUS", "JIL SANDER", "JUNJI",
            "JIMMY CHOO", "JORDAN", "JUNYA WATANABE", "KENZO", "LANVIN BLANC",
            "LANVIN", "LEMAIRE", "LOEWE", "LORO PIANA", "LOUBOUTIN", "LOUIS VUITTON",
            "MACKAGE", "MAISON MARGIELA", "MAISON MIHARA YASUHIRO", "MANOLO BLAHNIK",
            "MARNI", "MARTINE ROSE", "MAX MARA", "MAISON KITSUNE", "MIU MIU",
            "MONCLER", "MOOSE KNUCKLES", "NEW BALANCE", "NIKE", "OFF WHITE",
            "OMEGA", "PHILIPP PLEIN", "PANERAI", "PARAJUMPERS", "PALM ANGELS",
            "PALACE", "PATEK PHILIPPE", "PRADA", "PIAGET", "POLORALPHLAUREN",
            "RAY BAN", "RHUDE", "RICK OWENS", "RIMOWA", "ROGER VIVIER", "ROLEX",
            "SACAI", "SUPREME", "SAINT LAURENT", "SALOMON", "STUSSY", "STONE ISLAND",
            "TAG HEUER", "THE NORTH FACE", "THOM BROWNE", "TIFFANY & CO", "TOM FORD",
            "TUDOR", "UMA WANG", "VACHERON CONSTANTIN", "VALENTINO", "VETEMENTS",
            "VANCLEEF", "VERSACE", "WOOYOUNGMI", "YEEZY", "ZEGNA", "OTHERS",
        ]

        avail_1st_categories = [
            "상의", "아우터", "하의", "가방", "신발", "지갑", "시계", "패션잡화", "액세서리", "벨트"
        ]

        avail_2nd_categories = [
            "반팔 티셔츠", "긴팔 티셔츠", "니트/가디건", "맨투맨", "후드", "원피스", "셔츠", "드레스",
            "슬리브리스", "셋업", "기타 상의", "집업", "자켓", "패딩", "레더", "코트", "기타 아우터",
            "팬츠", "쇼츠", "트레이닝 팬츠", "데님", "스커트", "기타 하의", "미니백", "백팩", "숄더백",
            "토트백", "크로스백", "클러치", "캐리어", "핸드백", "더플백", "버킷백", "기타 가방",
            "스니커즈", "샌들/슬리퍼", "플랫", "로퍼", "더비/레이스업", "힐/펌프스", "부츠", "기타 신발",
            "반지갑", "카드지갑", "지퍼장지갑", "중/장지갑", "여권지갑", "WOC", "기타 지갑",
            "메탈", "가죽", "우레탄", "머플러/스카프", "아이웨어", "넥타이", "모자", "헤어액세서리",
            "기타 잡화", "반지", "목걸이", "팔찌", "귀걸이", "키링", "브로치", "기타 ACC"
        ]

        for index, data in enumerate(final):
            if data['브랜드'] not in avail_brands:
                final[index]['브랜드'] = ""
            if data['2차'] not in avail_1st_categories:
                final[index]['2차'] = ""
            if data['3차'] not in avail_2nd_categories:
                final[index]['3차'] = ""

        df = pd.DataFrame(final)

        # 각 사이트별로 엑셀 파일 생성
        for site_name, _, _ in category_data:
            site_data = df[df['거래처'] == site_name]
            if not site_data.empty:
                wb = openpyxl.Workbook()
                ws = wb.active

                ws.append(list(site_data.columns))
                alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                font = Font(name='Arial')

                for cell in ws[1]:
                    cell.alignment = alignment
                    cell.font = font

                for idx, row in enumerate(dataframe_to_rows(site_data, index=False, header=False)):
                    img_path = row[4]
                    if os.path.exists(img_path) and img_path != "":
                        img = OpenpyxlImage(img_path)
                        img.width, img.height = 80, 80
                        ws.add_image(img, f"E{idx + 2}")
                        ws.row_dimensions[idx + 2].height = 65
                        row[4] = ""

                    ws.append(row)

                    for cell in ws[idx + 2]:
                        cell.alignment = alignment
                        cell.font = font

                # 열 너비 설정
                ws.column_dimensions['B'].width = 18
                ws.column_dimensions['E'].width = 12
                ws.column_dimensions['F'].width = 8.25
                ws.column_dimensions['G'].width = 11.25
                ws.column_dimensions['H'].width = 12.75
                ws.column_dimensions['K'].width = 12.75
                ws.column_dimensions['L'].width = 15
                ws.column_dimensions['N'].width = 39
                ws.column_dimensions['O'].width = 22.5
                ws.column_dimensions['P'].width = 12
                ws.column_dimensions['Q'].width = 12
                ws.column_dimensions['R'].width = 12
                ws.column_dimensions['S'].width = 12
                ws.column_dimensions['T'].width = 12
                ws.column_dimensions['U'].width = 12
                ws.column_dimensions['V'].width = 12
                ws.column_dimensions['W'].width = 12
                ws.column_dimensions['X'].width = 12
                ws.column_dimensions['Y'].width = 20
                ws.column_dimensions['Z'].width = 20
                ws.column_dimensions['AA'].width = 12
                ws.column_dimensions['AB'].width = 12
                ws.column_dimensions['AC'].width = 12
                ws.column_dimensions['AD'].width = 12

                count = len(site_data)
                excel_filename = f"결과_{site_name}_{count}개_{timestamp}.xlsx"
                wb.save(excel_filename)

        input("\n작업 완료! 엔터를 눌러 종료하세요 : ")

    except Exception as e:
        # 치명적 오류 발생 시 로깅
        logger.critical("프로그램 실행 중 치명적 오류 발생", exc_info=True)
    finally:
        # stdout 복구
        sys.stdout = sys.__stdout__
